# import openpyxl
# import os
#
#
# def get_data_by_testid(test_id):
#
#     # Dynamically find the Excel file in the project folder
#     base_dir = os.path.dirname(os.path.dirname(__file__))
#     file_path = os.path.join(base_dir, "test_data", "flight_data.xlsx")
#     workbook = openpyxl.load_workbook(file_path)
#     sheet = workbook["Sheet1"]
#
#     # Loop through rows (skipping the header) to find the matching TestID
#     for row in sheet.iter_rows(min_row=2, values_only=True):
#         if row[0] == test_id:
#             return {
#                 "source": row[1],
#                 "destination": row[2]
#             }
#
#     # ASSERTION 1: Fail the test immediately if the TestID isn't found in Excel
#     assert False, f"CRITICAL ERROR: Test ID '{test_id}' was not found in the Excel file!"

import openpyxl
import os

def get_data_by_testid(test_id):

    # Dynamically find the Excel file in the project folder
    base_dir = os.path.dirname(os.path.dirname(__file__))
    file_path = os.path.join(base_dir, "test_data", "flight_data.xlsx")
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook["Sheet1"]

    # Loop through rows (skipping the header) to find the matching TestID
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == test_id:
            return {
                "source": row[1],
                "destination": row[2],
                "first_name": row[3],
                "last_name": row[4],
                "mobile": row[5],
                "email": row[6]
            }

    # ASSERTION 1: Fail the test immediately if the TestID isn't found in Excel
    assert False, f"CRITICAL ERROR: Test ID '{test_id}' was not found in the Excel file!"